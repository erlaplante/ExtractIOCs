# Example to extract IOCs from MDATP data that have been exported to Excel
# Data schema reference:
# https://docs.microsoft.com/en-us/windows/security/threat-protection/microsoft-defender-atp/advanced-hunting-schema-reference

import openpyxl
from collections import defaultdict
from pathlib import Path

home = str(Path.home())
eventsPath = home + '\\Desktop\\PulledEvents.xlsx'
wb = openpyxl.load_workbook(eventsPath)

sheet = wb['Sheet1']

# Loop through each row for tagged IOC data and append to list
# Assuming tagged data begins on row 4 and the tagged column for potential IOCs is column 3
taggedRows = []
for i in range(4, sheet.max_row+1):
	if (sheet.cell(row=i, column=3).value == 'IOC'):
		taggedRows.append(i)

# Used to retrieve column index for associated column name
# Assuming column names are on row 3 
def getColIndex(colName):
	for i in range(1, sheet.max_column+1):
		if (sheet.cell(row=3, column=i).value == colName):
			return i

eventColIndex = getColIndex('EventType')


# Retrieving all DeviceNetworkEvents (IPs and Domains) rows
networkRows = []
for i in range(len(taggedRows)):
	if (sheet.cell(row=taggedRows[i], column=eventColIndex).value == 'DeviceNetworkEvents'):
		networkRows.append(taggedRows[i])

# Retrieving unique IP cell data
ipColIndex = getColIndex('RemoteIP')
ips = set() 
for i in range(len(networkRows)):
	if (sheet.cell(row=networkRows[i], column=ipColIndex).value != None):
		ips.add(sheet.cell(row=networkRows[i], column=ipColIndex).value)

# Retrieving unique Domain cell data
domainColIndex = getColIndex('RemoteUrl')
domains = set() 
for i in range(len(networkRows)):
	if (sheet.cell(row=networkRows[i], column=domainColIndex).value != None):
		domains.add(sheet.cell(row=networkRows[i], column=domainColIndex).value)


# Retreiving all tagged DeviceFileEvent rows
fileSha1Index = getColIndex('SHA1')
fileNameIndex = getColIndex('FileName')

cfRows = []
for i in range(len(taggedRows)):
	if (sheet.cell(row=taggedRows[i], column=eventColIndex).value == 'DeviceFileEvents'):
		cfRows.append(taggedRows[i])

# Dictionary for both DeviceFileEvent and DeviceProcessEvent cell data (names and SHA1 hashes)
# Stores unique hashes and list of unique (set) names for each hash
hashes = defaultdict(set)

# Add SHA1 and DeviceFileEvents to hashes dictionary, if SHA1 is already in the dictionary add FileName to set
for i in range(len(cfRows)):
	curSHA1 = (sheet.cell(row=cfRows[i], column=fileSha1Index).value) 
	curFileName = (sheet.cell(row=cfRows[i], column=fileNameIndex).value) 
	hashes[curSHA1].add(curFileName)	


# Retreiving all tagged DeviceProcessEvents rows
# Assumes extracted DeviceProcessEvents columns are named as follows to be distinct from DeviceFileEvents
cpSha1Index = getColIndex('cpSHA1')
cpNameIndex = getColIndex('cpFileName')

cpRows = []
for i in range(len(taggedRows)):
	if (sheet.cell(row=taggedRows[i], column=eventColIndex).value == 'DeviceProcessEvents'):
		cpRows.append(taggedRows[i])

# Add SHA1 and DeviceProcess FileName to hashes dictionary, if SHA1 is already in the dictionary add create process name to set
for i in range(len(cpRows)):
	curCpSha1 = (sheet.cell(row=cpRows[i], column=cpSha1Index).value)
	curCpName = (sheet.cell(row=cpRows[i], column=cpNameIndex).value)
	hashes[curCpSha1].add(curCpName)


print(ips)
print(domains)
print(hashes)
